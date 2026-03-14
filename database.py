
import sqlite3
import openpyxl
import datetime

class Database:
    def __init__(self, db_name):
        self.conn = sqlite3.connect(db_name)
        self.conn.row_factory = sqlite3.Row
        self.create_tables()
        self.import_all_data()

    def create_tables(self):
        cursor = self.conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS manufacturers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                article TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                unit TEXT,
                price REAL,
                category_id INTEGER,
                manufacturer_id INTEGER,
                supplier_id INTEGER,
                discount REAL DEFAULT 0,
                quantity INTEGER DEFAULT 0,
                description TEXT,
                photo TEXT,
                FOREIGN KEY (category_id) REFERENCES categories(id),
                FOREIGN KEY (manufacturer_id) REFERENCES manufacturers(id),
                FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role TEXT NOT NULL,
                full_name TEXT NOT NULL,
                login TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                address TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_date TEXT,
                delivery_date TEXT,
                address_id INTEGER,
                user_id INTEGER,
                pickup_code TEXT,
                status TEXT,
                FOREIGN KEY (address_id) REFERENCES addresses(id),
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                product_article TEXT,
                quantity INTEGER,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (product_article) REFERENCES products(article)
            )
        """)
        self.conn.commit()

    def import_all_data(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] > 0:
            return
        self.import_users()
        self.import_addresses()
        self.import_products()
        self.import_orders()

    def import_users(self):
        wb = openpyxl.load_workbook("user_import.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            role, full_name, login, password = row
            if role and full_name and login and password:
                cursor.execute(
                    "INSERT INTO users (role, full_name, login, password) VALUES (?, ?, ?, ?)",
                    (role.strip(), full_name.strip(), login.strip(), password.strip()),
                )
        self.conn.commit()
        wb.close()

    def import_addresses(self):
        wb = openpyxl.load_workbook("Пункты выдачи_import.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()
        for row in sheet.iter_rows(min_row=1, values_only=True):
            address = row[0]
            if address:
                cursor.execute("INSERT INTO addresses (address) VALUES (?)", (address.strip(),))
        self.conn.commit()
        wb.close()

    def import_products(self):
        wb = openpyxl.load_workbook("Tovar.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()

        cat_cache, man_cache, sup_cache = {}, {}, {}

        def get_or_create(table, name, cache):
            if name in cache:
                return cache[name]
                
            queries = {
                "categories": ("SELECT id FROM categories WHERE name = ?", "INSERT INTO categories (name) VALUES (?)"),
                "manufacturers": ("SELECT id FROM manufacturers WHERE name = ?", "INSERT INTO manufacturers (name) VALUES (?)"),
                "suppliers": ("SELECT id FROM suppliers WHERE name = ?", "INSERT INTO suppliers (name) VALUES (?)")
            }
            sel_q, ins_q = queries[table]
            
            cursor.execute(sel_q, (name,))
            row = cursor.fetchone()
            if row:
                cache[name] = row[0]
            else:
                cursor.execute(ins_q, (name,))
                cache[name] = cursor.lastrowid
            return cache[name]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (article, name, unit, price, supplier, manufacturer, 
             category, discount, quantity, description, photo) = row
            if not article:
                continue
                
            cat_id = get_or_create("categories", category.strip(), cat_cache) if category else None
            man_id = get_or_create("manufacturers", manufacturer.strip(), man_cache) if manufacturer else None
            sup_id = get_or_create("suppliers", supplier.strip(), sup_cache) if supplier else None

            cursor.execute(
                """
                INSERT INTO products (article, name, unit, price, category_id, manufacturer_id, supplier_id, discount, quantity, description, photo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (article.strip(), name.strip(), unit, price, cat_id, man_id, sup_id, discount, quantity, description, photo),
            )
        self.conn.commit()
        wb.close()

    def _parse_date(self, date_str):
        if not date_str:
            return None
        date_str = str(date_str).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        return None

    def _process_order_items(self, cursor, order_num, items_str):
        if not items_str:
            return
        parts = [p.strip() for p in items_str.split(",")]
        for i in range(0, len(parts) - 1, 2):
            article = parts[i]
            try:
                qty = int(parts[i + 1])
            except ValueError:
                qty = 1
            cursor.execute(
                "INSERT INTO order_items (order_id, product_article, quantity) VALUES (?, ?, ?)",
                (order_num, article, qty),
            )

    def import_orders(self):
        wb = openpyxl.load_workbook("Заказ_import.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()

        cursor.execute("SELECT id, full_name FROM users")
        user_by_name = {u["full_name"]: u["id"] for u in cursor.fetchall()}

        cursor.execute("SELECT id, address FROM addresses ORDER BY id")
        address_by_index = {idx + 1: addr["id"] for idx, addr in enumerate(cursor.fetchall())}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = row[:8]
            (order_num, items_str, order_date, delivery_date, 
             address_num, client_fio, pickup_code, status) = row
            
            if not order_num:
                continue

            order_date_parsed = self._parse_date(order_date)
            delivery_date_parsed = self._parse_date(delivery_date)

            address_id = None
            if address_num and isinstance(address_num, (int, str)) and str(address_num).isdigit():
                idx = int(address_num)
                if idx in address_by_index:
                    address_id = address_by_index[idx]

            user_id = user_by_name.get(client_fio.strip() if client_fio else None)

            cursor.execute(
                """
                INSERT INTO orders (id, order_date, delivery_date, address_id, user_id, pickup_code, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
                (order_num, order_date_parsed, delivery_date_parsed, address_id, user_id, str(pickup_code).strip(), status.strip() if status else ""),
            )

            self._process_order_items(cursor, order_num, items_str)
            
        self.conn.commit()
        wb.close()
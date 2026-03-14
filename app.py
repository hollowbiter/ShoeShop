import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import datetime
import re

import openpyxl
from PIL import Image, ImageTk

# Конфигурация
DB_NAME = "shoe_shop.db"
IMAGE_FOLDER = "product_images"
DEFAULT_IMAGE = "picture.png"
ICON_FILE = "Icon.ico"
LOGO_FILE = "Icon.png"
MAX_IMAGE_SIZE = (300, 200)

# Цвета
COLOR_MAIN_BG = "#FFFFFF"
COLOR_EXTRA_BG = "#7FFF00"
COLOR_ACCENT = "#00FA9A"
COLOR_DISCOUNT_HIGH = "#2E8B57"
COLOR_OUT_OF_STOCK = "lightblue"

# Константы для устранения дублирования литералов (S1192)
FONT_DEFAULT = ("Times New Roman", 12)
FONT_TITLE = ("Times New Roman", 16)
FONT_ENTRY = ("Times New Roman", 14)
FONT_SMALL = ("Times New Roman", 10)
FONT_SMALL_BOLD = ("Times New Roman", 10, "bold")
FONT_BOLD = ("Times New Roman", 12, "bold")
TITLE_ERROR = "Ошибка"
ROLE_ADMIN = "Администратор"
ROLE_MANAGER = "Менеджер"
VALUE_ALL_SUPPLIERS = "Все поставщики"
VALUE_ALL_CATEGORIES = "Все категории"
STATE_NORMAL = "normal"
STATE_READONLY = "readonly"
TEXT_START = "1.0"


# Поле ввода с контекстным меню
class EntryWithContextMenu(tk.Entry):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.context_menu = tk.Menu(self, tearoff=0)  
    def show_context_menu(self, event):
        self.context_menu.tk_popup(event.x_root, event.y_root)


# Класс базы данных
class Database:
    def __init__(self, db_name):
        self.conn = sqlite3.connect(db_name)
        self.conn.row_factory = sqlite3.Row
        self.create_tables()
        self.import_all_data()

    def create_tables(self):
        cursor = self.conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS manufacturers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                article TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                unit TEXT,
                price REAL,
                category_id INTEGER,
                manufacturer_id INTEGER,
                supplier_id INTEGER,
                discount REAL DEFAULT 0,
                quantity INTEGER DEFAULT 0,
                description TEXT,
                photo TEXT,
                FOREIGN KEY (category_id) REFERENCES categories(id),
                FOREIGN KEY (manufacturer_id) REFERENCES manufacturers(id),
                FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role TEXT NOT NULL,
                full_name TEXT NOT NULL,
                login TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                address TEXT UNIQUE NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_date TEXT,
                delivery_date TEXT,
                address_id INTEGER,
                user_id INTEGER,
                pickup_code TEXT,
                status TEXT,
                FOREIGN KEY (address_id) REFERENCES addresses(id),
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                product_article TEXT,
                quantity INTEGER,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE,
                FOREIGN KEY (product_article) REFERENCES products(article)
            )
        """)
        self.conn.commit()

    def import_all_data(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] > 0:
            return
        self.import_users()
        self.import_addresses()
        self.import_products()
        self.import_orders()

    def import_users(self):
        wb = openpyxl.load_workbook("user_import.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            role, full_name, login, password = row
            if role and full_name and login and password:
                cursor.execute(
                    "INSERT INTO users (role, full_name, login, password) VALUES (?, ?, ?, ?)",
                    (role.strip(), full_name.strip(), login.strip(), password.strip()),
                )
        self.conn.commit()
        wb.close()

    def import_addresses(self):
        wb = openpyxl.load_workbook("Пункты выдачи_import.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()
        for row in sheet.iter_rows(min_row=1, values_only=True):
            address = row[0]
            if address:
                cursor.execute("INSERT INTO addresses (address) VALUES (?)", (address.strip(),))
        self.conn.commit()
        wb.close()

    def import_products(self):
        wb = openpyxl.load_workbook("Tovar.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()

        cat_cache, man_cache, sup_cache = {}, {}, {}

        def get_or_create(table, name, cache):
            if name in cache:
                return cache[name]
                
            queries = {
                "categories": ("SELECT id FROM categories WHERE name = ?", "INSERT INTO categories (name) VALUES (?)"),
                "manufacturers": ("SELECT id FROM manufacturers WHERE name = ?", "INSERT INTO manufacturers (name) VALUES (?)"),
                "suppliers": ("SELECT id FROM suppliers WHERE name = ?", "INSERT INTO suppliers (name) VALUES (?)")
            }
            sel_q, ins_q = queries[table]
            
            cursor.execute(sel_q, (name,))
            row = cursor.fetchone()
            if row:
                cache[name] = row[0]
            else:
                cursor.execute(ins_q, (name,))
                cache[name] = cursor.lastrowid
            return cache[name]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (article, name, unit, price, supplier, manufacturer, 
             category, discount, quantity, description, photo) = row
            if not article:
                continue
                
            cat_id = get_or_create("categories", category.strip(), cat_cache) if category else None
            man_id = get_or_create("manufacturers", manufacturer.strip(), man_cache) if manufacturer else None
            sup_id = get_or_create("suppliers", supplier.strip(), sup_cache) if supplier else None

            cursor.execute(
                """
                INSERT INTO products (article, name, unit, price, category_id, manufacturer_id, supplier_id, discount, quantity, description, photo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (article.strip(), name.strip(), unit, price, cat_id, man_id, sup_id, discount, quantity, description, photo),
            )
        self.conn.commit()
        wb.close()

    # Хелпер для снижения сложности import_orders
    def _parse_date(self, date_str):
        if not date_str:
            return None
        date_str = str(date_str).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d"):
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        return None

    # Хелпер для снижения сложности import_orders
    def _process_order_items(self, cursor, order_num, items_str):
        if not items_str:
            return
        parts = [p.strip() for p in items_str.split(",")]
        for i in range(0, len(parts) - 1, 2):
            article = parts[i]
            try:
                qty = int(parts[i + 1])
            except ValueError:
                qty = 1
            cursor.execute(
                "INSERT INTO order_items (order_id, product_article, quantity) VALUES (?, ?, ?)",
                (order_num, article, qty),
            )

    def import_orders(self):
        wb = openpyxl.load_workbook("Заказ_import.xlsx", data_only=True)
        sheet = wb.active
        cursor = self.conn.cursor()

        cursor.execute("SELECT id, full_name FROM users")
        user_by_name = {u["full_name"]: u["id"] for u in cursor.fetchall()}

        cursor.execute("SELECT id, address FROM addresses ORDER BY id")
        address_by_index = {idx + 1: addr["id"] for idx, addr in enumerate(cursor.fetchall())}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = row[:8]
            (order_num, items_str, order_date, delivery_date, 
             address_num, client_fio, pickup_code, status) = row
            
            if not order_num:
                continue

            order_date_parsed = self._parse_date(order_date)
            delivery_date_parsed = self._parse_date(delivery_date)

            address_id = None
            if address_num and isinstance(address_num, (int, str)) and str(address_num).isdigit():
                idx = int(address_num)
                if idx in address_by_index:
                    address_id = address_by_index[idx]

            user_id = user_by_name.get(client_fio.strip() if client_fio else None)

            cursor.execute(
                """
                INSERT INTO orders (id, order_date, delivery_date, address_id, user_id, pickup_code, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
                (order_num, order_date_parsed, delivery_date_parsed, address_id, user_id, str(pickup_code).strip(), status.strip() if status else ""),
            )

            self._process_order_items(cursor, order_num, items_str)
            
        self.conn.commit()
        wb.close()


# Главное приложение
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ООО Обувь - Вход")
        self.geometry("1200x800")
        self.resizable(True, True)
        if os.path.exists(ICON_FILE):
            self.iconbitmap(ICON_FILE)
        self.db = Database(DB_NAME)
        self.current_user = None
        if not os.path.exists(IMAGE_FOLDER):
            os.makedirs(IMAGE_FOLDER)
        self.show_login()

    def show_login(self):
        self.title("ООО Обувь - Вход")
        self.clear_window()
        LoginWindow(self)

    def show_products(self):
        self.title("ООО Обувь - Товары")
        self.clear_window()
        ProductsWindow(self)

    def show_orders(self):
        self.title("ООО Обувь - Заказы")
        self.clear_window()
        OrdersWindow(self)

    def clear_window(self):
        for widget in self.winfo_children():
            widget.destroy()


# Окно входа
class LoginWindow(tk.Frame):
    def __init__(self, app):
        super().__init__(app, bg=COLOR_MAIN_BG)
        self.app = app
        self.pack(fill=tk.BOTH, expand=True)

        if os.path.exists(LOGO_FILE):
            logo_img = Image.open(LOGO_FILE)
            logo_img = logo_img.resize((150, 150), Image.Resampling.LANCZOS)
            self.logo = ImageTk.PhotoImage(logo_img)
            tk.Label(self, image=self.logo, bg=COLOR_MAIN_BG).pack(pady=10)

        tk.Label(self, text="Вход в систему", font=FONT_TITLE, bg=COLOR_MAIN_BG).pack(pady=5)

        frame = tk.Frame(self, bg=COLOR_MAIN_BG)
        frame.pack(pady=10)

        tk.Label(frame, text="Логин:", bg=COLOR_MAIN_BG, font=FONT_DEFAULT).grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.entry_login = EntryWithContextMenu(frame, font=FONT_ENTRY, width=25, bd=3, relief=tk.RIDGE)
        self.entry_login.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(frame, text="Пароль:", bg=COLOR_MAIN_BG, font=FONT_DEFAULT).grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.entry_password = EntryWithContextMenu(frame, show="*", font=FONT_ENTRY, width=25, bd=3, relief=tk.RIDGE)
        self.entry_password.grid(row=1, column=1, padx=5, pady=5)

        btn_frame = tk.Frame(self, bg=COLOR_MAIN_BG)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Войти", command=self.login, bg=COLOR_ACCENT, 
          font=FONT_ENTRY, padx=15, pady=5).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Войти как гость", command=self.guest_login, bg=COLOR_EXTRA_BG,
          font=FONT_ENTRY, padx=15, pady=5).pack(side=tk.LEFT, padx=5)

    def login(self):
        login = self.entry_login.get().strip()
        password = self.entry_password.get().strip()
        cursor = self.app.db.conn.cursor()
        cursor.execute("SELECT * FROM users WHERE login=? AND password=?", (login, password))
        user = cursor.fetchone()
        if user:
            self.app.current_user = dict(user)
            self.app.show_products()
        else:
            messagebox.showerror(TITLE_ERROR, "Неверный логин или пароль")

    def guest_login(self):
        self.app.current_user = None
        self.app.show_products()


# Окно товаров
class ProductsWindow(tk.Frame):
    def __init__(self, app):
        super().__init__(app, bg=COLOR_MAIN_BG)
        self.app = app
        self.pack(fill=tk.BOTH, expand=True)

        top_frame = tk.Frame(self, bg=COLOR_EXTRA_BG, height=40)
        top_frame.pack(fill=tk.X, side=tk.TOP)
        top_frame.pack_propagate(False)

        if self.app.current_user:
            fio = self.app.current_user["full_name"]
            role = self.app.current_user["role"]
            text = f"{fio} ({role})"
        else:
            text = "Гость"
        tk.Label(top_frame, text=text, bg=COLOR_EXTRA_BG, font=FONT_DEFAULT).pack(side=tk.RIGHT, padx=10, pady=5)
        tk.Button(top_frame, text="Выход", command=self.logout, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=10, pady=5)

        has_access = self.app.current_user and self.app.current_user["role"] in (ROLE_MANAGER, ROLE_ADMIN)
        if has_access:
            tk.Button(top_frame, text="Заказы", command=self.app.show_orders, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=10, pady=5)
            self.create_filter_panel()

        if self.app.current_user and self.app.current_user["role"] == ROLE_ADMIN:
            tk.Button(top_frame, text="Добавить товар", command=self.add_product, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=10, pady=5)

        self.canvas = tk.Canvas(self, bg=COLOR_MAIN_BG, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=COLOR_MAIN_BG)

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.load_products()

    def create_filter_panel(self):
        filter_frame = tk.Frame(self, bg=COLOR_MAIN_BG)
        filter_frame.pack(fill=tk.X, pady=5, padx=10)

        # Поиск
        tk.Label(filter_frame, text="Поиск:", bg=COLOR_MAIN_BG, font=FONT_SMALL).pack(side=tk.LEFT, padx=(5,2))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.load_products())
        search_entry = tk.Entry(filter_frame, textvariable=self.search_var, font=FONT_SMALL, width=25, bd=2, relief=tk.SUNKEN)
        search_entry.pack(side=tk.LEFT, padx=(0,15))

        # Сортировка
        tk.Label(filter_frame, text="Сорт.:", bg=COLOR_MAIN_BG, font=FONT_SMALL).pack(side=tk.LEFT, padx=(5,2))
        self.sort_var = tk.StringVar(value="Нет")
        self.sort_var.trace('w', lambda *args: self.load_products())
        sort_combo = ttk.Combobox(filter_frame, textvariable=self.sort_var, values=["Нет", "По возрастанию", "По убыванию"], state=STATE_READONLY, width=15)
        sort_combo.pack(side=tk.LEFT, padx=(0,15))

        # Поставщик
        tk.Label(filter_frame, text="Поставщик:", bg=COLOR_MAIN_BG, font=FONT_SMALL).pack(side=tk.LEFT, padx=(5,2))
        self.supplier_var = tk.StringVar(value=VALUE_ALL_SUPPLIERS)
        self.supplier_var.trace('w', lambda *args: self.load_products())
        cursor = self.app.db.conn.cursor()
        cursor.execute("SELECT name FROM suppliers ORDER BY name")
        suppliers = [VALUE_ALL_SUPPLIERS] + [row['name'] for row in cursor.fetchall()]
        supplier_combo = ttk.Combobox(filter_frame, textvariable=self.supplier_var, values=suppliers, state=STATE_READONLY, width=18)
        supplier_combo.pack(side=tk.LEFT, padx=(0,15))

        # Категория
        tk.Label(filter_frame, text="Категория:", bg=COLOR_MAIN_BG, font=FONT_SMALL).pack(side=tk.LEFT, padx=(5,2))
        self.category_var = tk.StringVar(value=VALUE_ALL_CATEGORIES)
        self.category_var.trace('w', lambda *args: self.load_products())
        cursor.execute("SELECT name FROM categories ORDER BY name")
        categories = [VALUE_ALL_CATEGORIES] + [row['name'] for row in cursor.fetchall()]
        category_combo = ttk.Combobox(filter_frame, textvariable=self.category_var, values=categories, state=STATE_READONLY, width=15)
        category_combo.pack(side=tk.LEFT, padx=(0,5))

    # Хелпер для снижения сложности load_products
    def _build_product_query(self):
        query = '''
            SELECT p.*, cat.name as category_name, man.name as manufacturer_name, sup.name as supplier_name
            FROM products p
            LEFT JOIN categories cat ON p.category_id = cat.id
            LEFT JOIN manufacturers man ON p.manufacturer_id = man.id
            LEFT JOIN suppliers sup ON p.supplier_id = sup.id
        '''
        conditions, params = [], []

        if hasattr(self, 'search_var') and self.search_var.get().strip():
            search = self.search_var.get().strip()
            conditions.append('(p.article LIKE ? OR p.name LIKE ? OR p.description LIKE ? OR cat.name LIKE ? OR man.name LIKE ? OR sup.name LIKE ?)')
            params.extend([f'%{search}%'] * 6)

        if hasattr(self, 'supplier_var') and self.supplier_var.get() != VALUE_ALL_SUPPLIERS:
            conditions.append("sup.name = ?")
            params.append(self.supplier_var.get())

        if hasattr(self, 'category_var') and self.category_var.get() != VALUE_ALL_CATEGORIES:
            conditions.append("cat.name = ?")
            params.append(self.category_var.get())

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        if hasattr(self, 'sort_var'):
            if self.sort_var.get() == "По возрастанию":
                query += " ORDER BY p.quantity ASC"
            elif self.sort_var.get() == "По убыванию":
                query += " ORDER BY p.quantity DESC"

        return query, params

    def load_products(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        cursor = self.app.db.conn.cursor()
        query, params = self._build_product_query()

        try:
            cursor.execute(query, params)
            products = cursor.fetchall()
        except Exception as e:
            print("Ошибка при выполнении запроса:", e)
            products = []

        for prod in products:
            self.create_product_card(prod)

        self.scrollable_frame.update_idletasks()
        self.canvas.update_idletasks()
        bbox = self.canvas.bbox("all")
        if bbox and bbox[3] <= self.canvas.winfo_height():
            self.canvas.configure(scrollregion=bbox)
            self.scrollbar.pack_forget()
        else:
            self.canvas.configure(scrollregion=bbox)
            self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def create_product_card(self, prod):
        card = tk.Frame(self.scrollable_frame, bg=COLOR_MAIN_BG, relief=tk.RAISED, bd=1)
        card.pack(fill=tk.X, padx=10, pady=5, ipadx=5, ipady=5)

        bg_color = COLOR_MAIN_BG
        if prod["quantity"] == 0:
            bg_color = COLOR_OUT_OF_STOCK
        elif prod["discount"] > 15:
            bg_color = COLOR_DISCOUNT_HIGH
        card.config(bg=bg_color)
        
        img_frame = tk.Frame(card, bg=COLOR_MAIN_BG, width=120, height=120)
        img_frame.pack(side=tk.LEFT, padx=5, pady=5)
        img_frame.pack_propagate(False)

        img_path = prod['photo']
        if img_path and os.path.exists(img_path):
            pil_img = Image.open(img_path)
        else:
            pil_img = Image.open(DEFAULT_IMAGE)
        
        pil_img.thumbnail((120, 120), Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(pil_img)
        img_label = tk.Label(img_frame, image=photo, bg=COLOR_MAIN_BG)
        img_label.image = photo
        img_label.pack(expand=True, fill=tk.BOTH)

        text_frame = tk.Frame(card, bg=bg_color)
        text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        tk.Label(text_frame, text=f"{prod['category_name'] or ''} | {prod['name'] or ''}", font=FONT_BOLD, bg=bg_color, anchor=tk.W).pack(anchor=tk.W, fill=tk.X)

        desc = prod["description"] or ""
        if len(desc) > 100:
            desc = desc[:100] + "..."
        tk.Label(text_frame, text=f"Описание товара: {desc}", bg=bg_color, font=FONT_SMALL, wraplength=400, justify=tk.LEFT, anchor=tk.W).pack(anchor=tk.W, fill=tk.X)
        tk.Label(text_frame, text=f"Производитель: {prod['manufacturer_name'] or ''}", bg=bg_color, font=FONT_SMALL, anchor=tk.W).pack(anchor=tk.W, fill=tk.X)
        tk.Label(text_frame, text=f"Поставщик: {prod['supplier_name'] or ''}", bg=bg_color, font=FONT_SMALL, anchor=tk.W).pack(anchor=tk.W, fill=tk.X)

        original_price = prod["price"]
        discount = prod["discount"] or 0
        final_price = original_price * (1 - discount / 100)

        price_frame = tk.Frame(text_frame, bg=bg_color)
        price_frame.pack(anchor=tk.W, fill=tk.X, pady=2)

        if discount > 0:
            tk.Label(price_frame, text=f"{original_price:.2f} руб.", fg="red", bg=bg_color, font=("Times New Roman", 10, "overstrike")).pack(side=tk.LEFT, padx=(0, 5))
            tk.Label(price_frame, text=f"{final_price:.2f} руб.", fg="black", bg=bg_color, font=FONT_SMALL_BOLD).pack(side=tk.LEFT)
        else:
            tk.Label(price_frame, text=f"{original_price:.2f} руб.", font=FONT_SMALL, bg=bg_color).pack(side=tk.LEFT)

        tk.Label(text_frame, text=f"Единица измерения: {prod['unit'] or ''}", bg=bg_color, font=FONT_SMALL, anchor=tk.W).pack(anchor=tk.W, fill=tk.X)
        tk.Label(text_frame, text=f"Количество на складе: {prod['quantity']}", bg=bg_color, font=FONT_SMALL, anchor=tk.W).pack(anchor=tk.W, fill=tk.X)

        if discount > 0:
            disc_frame = tk.Frame(card, bg=COLOR_MAIN_BG, width=110, height=110, relief=tk.RIDGE, bd=2)
            disc_frame.pack(side=tk.RIGHT, padx=5, pady=5)
            disc_frame.pack_propagate(False)
            inner_frame = tk.Frame(disc_frame, bg=COLOR_MAIN_BG)
            inner_frame.pack(expand=True)
            tk.Label(inner_frame, text="Скидка", bg=COLOR_MAIN_BG, font=FONT_SMALL_BOLD).pack(anchor="center")
            tk.Label(inner_frame, text=f"{int(discount)}%", bg=COLOR_MAIN_BG, font=("Times New Roman", 28, "bold"), fg="red").pack(anchor="center")

        if self.app.current_user and self.app.current_user["role"] == ROLE_ADMIN:
            btn_frame = tk.Frame(text_frame, bg=bg_color)
            btn_frame.pack(anchor="e", pady=2, fill=tk.X)
            tk.Button(btn_frame, text="Редактировать", command=lambda p=prod: self.edit_product(p["article"]), bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=2)
            tk.Button(btn_frame, text="Удалить", command=lambda p=prod: self.delete_product(p["article"]), bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=2)

    def logout(self):
        self.app.current_user = None
        self.app.show_login()

    def add_product(self):
        ProductEditWindow(self.app, self)

    def edit_product(self, article):
        ProductEditWindow(self.app, self, article)

    def delete_product(self, article):
        cursor = self.app.db.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM order_items WHERE product_article=?", (article,))
        if cursor.fetchone()[0] > 0:
            messagebox.showerror(TITLE_ERROR, "Нельзя удалить товар, который присутствует в заказах")
            return
        if messagebox.askyesno("Подтверждение", f"Удалить товар {article}?"):
            cursor.execute("DELETE FROM products WHERE article=?", (article,))
            self.app.db.conn.commit()
            self.load_products()


# Окно заказов
class OrdersWindow(tk.Frame):
    def __init__(self, app):
        super().__init__(app, bg=COLOR_MAIN_BG)
        self.app = app
        self.pack(fill=tk.BOTH, expand=True)

        top_frame = tk.Frame(self, bg=COLOR_EXTRA_BG, height=40)
        top_frame.pack(fill=tk.X, side=tk.TOP)
        top_frame.pack_propagate(False)

        user_text = "Гость"
        if self.app.current_user:
            user_text = f"{self.app.current_user['full_name']} ({self.app.current_user['role']})"
        
        tk.Label(top_frame, text=user_text, bg=COLOR_EXTRA_BG, font=FONT_DEFAULT).pack(side=tk.RIGHT, padx=10)
        tk.Button(top_frame, text="Назад к товарам", command=self.app.show_products, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=10, pady=5)
        
        if self.app.current_user and self.app.current_user["role"] == ROLE_ADMIN:
            tk.Button(top_frame, text="Добавить заказ", command=self.add_order, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=10, pady=5)

        self.canvas = tk.Canvas(self, bg=COLOR_MAIN_BG, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=COLOR_MAIN_BG)

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.load_orders()

    def load_orders(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        try:
            cursor = self.app.db.conn.cursor()
            cursor.execute("""
                SELECT o.id, o.order_date, o.delivery_date, a.address, u.full_name, o.pickup_code, o.status
                FROM orders o
                LEFT JOIN addresses a ON o.address_id = a.id
                LEFT JOIN users u ON o.user_id = u.id
                ORDER BY o.id DESC
            """)
            
            rows = cursor.fetchall()
            if not rows:
                tk.Label(self.scrollable_frame, text="Заказов пока нет", bg=COLOR_MAIN_BG, font=FONT_ENTRY).pack(pady=50)
                return

            for row in rows:
                self.create_order_card(row)
        except Exception as e:
            tk.Label(self.scrollable_frame, text=f"Ошибка базы данных: {e}", fg="red", bg=COLOR_MAIN_BG).pack()

    def create_order_card(self, row):
        card = tk.Frame(self.scrollable_frame, bg=COLOR_MAIN_BG, relief=tk.RIDGE, bd=2)
        card.pack(fill=tk.X, padx=20, pady=10, ipady=10)

        left_box = tk.Frame(card, bg=COLOR_MAIN_BG)
        left_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        tk.Label(left_box, text=f"Артикул заказа: {row[0]}", font=FONT_BOLD, bg=COLOR_MAIN_BG, anchor=tk.W).pack(fill=tk.X)
        tk.Label(left_box, text=f"Статус заказа: {row[6]}", font=("Times New Roman", 11), bg=COLOR_MAIN_BG, anchor=tk.W).pack(fill=tk.X)
        
        addr = row[3] if row[3] else "Адрес не указан"
        tk.Label(left_box, text=f"Адрес пункта выдачи: {addr}", font=("Times New Roman", 11), bg=COLOR_MAIN_BG, anchor=tk.W, wraplength=600).pack(fill=tk.X)
        tk.Label(left_box, text=f"Дата заказа: {row[1]}", font=("Times New Roman", 11), bg=COLOR_MAIN_BG, anchor=tk.W).pack(fill=tk.X)

        right_box = tk.Frame(card, bg=COLOR_MAIN_BG, relief=tk.SOLID, bd=1, width=150)
        right_box.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=5)
        right_box.pack_propagate(False)

        tk.Label(right_box, text="Дата доставки", font=FONT_SMALL, bg=COLOR_MAIN_BG).pack(pady=(5, 0))
        tk.Label(right_box, text=f"{row[2]}", font=("Times New Roman", 11, "bold"), bg=COLOR_MAIN_BG).pack(expand=True)

        if self.app.current_user and self.app.current_user["role"] == ROLE_ADMIN:
            btn_frame = tk.Frame(left_box, bg=COLOR_MAIN_BG)
            btn_frame.pack(anchor=tk.W, pady=(10, 0))
            order_id = row[0]
            tk.Button(btn_frame, text="Редактировать", command=lambda i=order_id: OrderEditWindow(self.app, self, i), bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=2)
            tk.Button(btn_frame, text="Удалить", command=lambda i=order_id: self.delete_order(i), bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=2)

    def delete_order(self, order_id):
        if messagebox.askyesno("Подтверждение", f"Удалить заказ №{order_id}?"):
            cursor = self.app.db.conn.cursor()
            cursor.execute("DELETE FROM orders WHERE id=?", (order_id,))
            self.app.db.conn.commit()
            self.load_orders()

    def add_order(self):
        OrderEditWindow(self.app, self)

    def logout(self):
        self.app.current_user = None
        self.app.show_login()


# Окно редактирования товара
class ProductEditWindow(tk.Toplevel):
    def __init__(self, app, parent_window, article=None):
        super().__init__(app)
        self.app = app
        self.parent = parent_window
        self.article = article
        self.title("Редактирование товара" if article else "Добавление товара")
        self.geometry("800x900")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.transient(app)
        self.grab_set()

        if hasattr(app, "edit_window") and app.edit_window and app.edit_window.winfo_exists():
            messagebox.showwarning("Предупреждение", "Окно редактирования уже открыто")
            self.destroy()
            return
        app.edit_window = self

        self.load_combobox_data()
        self.create_widgets()
        if article:
            self.load_product_data()

    def load_combobox_data(self):
        cursor = self.app.db.conn.cursor()
        cursor.execute("SELECT name FROM categories ORDER BY name")
        self.categories = [row["name"] for row in cursor.fetchall()]
        cursor.execute("SELECT name FROM manufacturers ORDER BY name")
        self.manufacturers = [row["name"] for row in cursor.fetchall()]
        cursor.execute("SELECT name FROM suppliers ORDER BY name")
        self.suppliers = [row["name"] for row in cursor.fetchall()]

    def create_widgets(self):
        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, text="Артикул:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.entry_article = tk.Entry(main_frame, state=STATE_READONLY if self.article else STATE_NORMAL)
        self.entry_article.grid(row=0, column=1, sticky="ew", pady=2)
        if not self.article:
            self.entry_article.insert(0, self.generate_article())

        tk.Label(main_frame, text="Наименование:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.entry_name = tk.Entry(main_frame)
        self.entry_name.grid(row=1, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Категория:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.combo_category = ttk.Combobox(main_frame, values=self.categories, state=STATE_NORMAL)
        self.combo_category.grid(row=2, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Производитель:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.combo_manufacturer = ttk.Combobox(main_frame, values=self.manufacturers, state=STATE_NORMAL)
        self.combo_manufacturer.grid(row=3, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Поставщик:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.combo_supplier = ttk.Combobox(main_frame, values=self.suppliers, state=STATE_NORMAL)
        self.combo_supplier.grid(row=4, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Цена:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.entry_price = tk.Entry(main_frame)
        self.entry_price.grid(row=5, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Единица измерения:").grid(row=6, column=0, sticky=tk.W, pady=2)
        self.entry_unit = tk.Entry(main_frame)
        self.entry_unit.grid(row=6, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Количество на складе:").grid(row=7, column=0, sticky=tk.W, pady=2)
        self.entry_quantity = tk.Entry(main_frame)
        self.entry_quantity.grid(row=7, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Скидка (%):").grid(row=8, column=0, sticky=tk.W, pady=2)
        self.entry_discount = tk.Entry(main_frame)
        self.entry_discount.grid(row=8, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Описание:").grid(row=9, column=0, sticky=tk.W, pady=2)
        self.text_description = tk.Text(main_frame, height=5, width=40)
        self.text_description.grid(row=9, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Фото:").grid(row=10, column=0, sticky=tk.W, pady=2)
        self.photo_path = tk.StringVar()
        tk.Entry(main_frame, textvariable=self.photo_path, state=STATE_READONLY).grid(row=10, column=1, sticky="ew", pady=2)
        tk.Button(main_frame, text="Выбрать файл", command=self.select_photo).grid(row=10, column=2, padx=5)

        self.preview_label = tk.Label(main_frame, text="Предпросмотр")
        self.preview_label.grid(row=11, column=1, pady=5)

        btn_frame = tk.Frame(main_frame)
        btn_frame.grid(row=12, column=0, columnspan=3, pady=10)
        tk.Button(btn_frame, text="Сохранить", command=self.save_product, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Отмена", command=self.on_close, bg=COLOR_EXTRA_BG).pack(side=tk.LEFT, padx=5)

        main_frame.columnconfigure(1, weight=1)

    def generate_article(self):
        cursor = self.app.db.conn.cursor()
        cursor.execute("SELECT article FROM products")
        existing = [row['article'] for row in cursor.fetchall()]
        max_num = 0
        for art in existing:
            match = re.search(r'[A-Za-z](\d{3})', art)
            if match:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num
        next_num = max_num + 1
        new_article = f"A{next_num:03d}T4"
        while new_article in existing:
            next_num += 1
            new_article = f"A{next_num:03d}T4"
        return new_article

    def load_product_data(self):
        cursor = self.app.db.conn.cursor()
        cursor.execute("""
            SELECT p.*, cat.name as category_name, man.name as manufacturer_name, sup.name as supplier_name
            FROM products p
            LEFT JOIN categories cat ON p.category_id = cat.id
            LEFT JOIN manufacturers man ON p.manufacturer_id = man.id
            LEFT JOIN suppliers sup ON p.supplier_id = sup.id
            WHERE p.article = ?
        """, (self.article,))
        prod = cursor.fetchone()
        if prod:
            self.entry_article.config(state=STATE_NORMAL)
            self.entry_article.delete(0, tk.END)
            self.entry_article.insert(0, prod["article"])
            self.entry_article.config(state=STATE_READONLY)
            self.entry_name.insert(0, prod["name"] or "")
            self.combo_category.set(prod["category_name"] or "")
            self.combo_manufacturer.set(prod["manufacturer_name"] or "")
            self.combo_supplier.set(prod["supplier_name"] or "")
            self.entry_price.insert(0, prod["price"] or "")
            self.entry_unit.insert(0, prod["unit"] or "")
            self.entry_quantity.insert(0, prod["quantity"] or "")
            self.entry_discount.insert(0, prod["discount"] or "")
            self.text_description.insert(TEXT_START, prod["description"] or "")
            self.photo_path.set(prod["photo"] or "")
            self.update_preview()

    def select_photo(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp")])
        if file_path:
            self.photo_path.set(file_path)
            self.update_preview()

    def update_preview(self):
        path = self.photo_path.get()
        if path and os.path.exists(path):
            img = Image.open(path)
            img.thumbnail((100, 100))
            photo = ImageTk.PhotoImage(img)
            self.preview_label.config(image=photo, text="")
            self.preview_label.image = photo
        else:
            self.preview_label.config(image="", text="Нет изображения")

    def save_product(self):
        article = self.entry_article.get().strip()
        name = self.entry_name.get().strip()
        category = self.combo_category.get().strip()
        manufacturer = self.combo_manufacturer.get().strip()
        supplier = self.combo_supplier.get().strip()
        price_str = self.entry_price.get().strip()
        unit = self.entry_unit.get().strip()
        qty_str = self.entry_quantity.get().strip()
        disc_str = self.entry_discount.get().strip()
        description = self.text_description.get(TEXT_START, tk.END).strip()
        photo_src = self.photo_path.get().strip()

        if not article or not name or not price_str:
            messagebox.showerror(TITLE_ERROR, "Заполните обязательные поля (артикул, наименование, цена)")
            return
            
        try:
            price = float(price_str)
            quantity = int(qty_str) if qty_str else 0
            discount = float(disc_str) if disc_str else 0
            if price < 0 or quantity < 0 or not (0 <= discount <= 100):
                raise ValueError
        except ValueError:
            messagebox.showerror(TITLE_ERROR, "Убедитесь, что цена и количество — положительные числа, а скидка от 0 до 100.")
            return

        final_photo_path = None
        if photo_src and os.path.exists(photo_src):
            ext = os.path.splitext(photo_src)[1]
            dest_filename = f"{article}{ext}"
            dest_path = os.path.join(IMAGE_FOLDER, dest_filename)
            if self.article:
                cursor = self.app.db.conn.cursor()
                cursor.execute("SELECT photo FROM products WHERE article=?", (self.article,))
                old_photo = cursor.fetchone()
                if old_photo and old_photo["photo"] and os.path.exists(old_photo["photo"]):
                    os.remove(old_photo["photo"])
            img = Image.open(photo_src)
            img.thumbnail(MAX_IMAGE_SIZE, Image.Resampling.LANCZOS)
            img.save(dest_path)
            final_photo_path = dest_path

        def get_or_create_id(table, name):
            if not name:
                return None
            cursor = self.app.db.conn.cursor()
            queries = {
                "categories": ("SELECT id FROM categories WHERE name=?", "INSERT INTO categories (name) VALUES (?)"),
                "manufacturers": ("SELECT id FROM manufacturers WHERE name=?", "INSERT INTO manufacturers (name) VALUES (?)"),
                "suppliers": ("SELECT id FROM suppliers WHERE name=?", "INSERT INTO suppliers (name) VALUES (?)")
            }
            sel_q, ins_q = queries[table]
            cursor.execute(sel_q, (name,))
            row = cursor.fetchone()
            if row:
                return row[0]
            cursor.execute(ins_q, (name,))
            return cursor.lastrowid

        cat_id = get_or_create_id("categories", category)
        man_id = get_or_create_id("manufacturers", manufacturer)
        sup_id = get_or_create_id("suppliers", supplier)

        cursor = self.app.db.conn.cursor()
        if self.article:
            cursor.execute("""
                UPDATE products SET name=?, unit=?, price=?, category_id=?, manufacturer_id=?, supplier_id=?, discount=?, quantity=?, description=?, photo=?
                WHERE article=?
            """, (name, unit, price, cat_id, man_id, sup_id, discount, quantity, description, final_photo_path, self.article))
        else:
            cursor.execute("""
                INSERT INTO products (article, name, unit, price, category_id, manufacturer_id, supplier_id, discount, quantity, description, photo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (article, name, unit, price, cat_id, man_id, sup_id, discount, quantity, description, final_photo_path))

        self.app.db.conn.commit()
        self.parent.load_products()
        self.on_close()

    def on_close(self):
        self.app.edit_window = None
        self.destroy()


# Окно редактирования заказа
class OrderEditWindow(tk.Toplevel):
    def __init__(self, app, parent_window, order_id=None):
        super().__init__(app)
        self.app = app
        self.parent = parent_window
        self.order_id = order_id
        self.title("Редактирование заказа" if order_id else "Добавление заказа")
        self.geometry("800x700")
        self.resizable(True, True)
        self.transient(app)
        self.grab_set()

        if hasattr(app, "order_edit_window") and app.order_edit_window and app.order_edit_window.winfo_exists():
            messagebox.showwarning("Предупреждение", "Окно редактирования заказа уже открыто")
            self.destroy()
            return
        app.order_edit_window = self

        self.load_combobox_data()
        self.create_widgets()
        if order_id:
            self.load_order_data()

    def load_combobox_data(self):
        cursor = self.app.db.conn.cursor()
        cursor.execute("SELECT id, address FROM addresses ORDER BY id")
        self.addresses = cursor.fetchall()
        cursor.execute("SELECT id, full_name FROM users ORDER BY full_name")
        self.users = cursor.fetchall()
        self.statuses = ["Новый", "Завершен", "Отменен"]

    def create_widgets(self):
        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, text="Номер заказа:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.entry_id = tk.Entry(main_frame, state=STATE_READONLY if self.order_id else STATE_NORMAL)
        self.entry_id.grid(row=0, column=1, sticky="ew", pady=2)
        if not self.order_id:
            cursor = self.app.db.conn.cursor()
            cursor.execute("SELECT MAX(id) FROM orders")
            max_id = cursor.fetchone()[0] or 0
            self.entry_id.insert(0, str(max_id + 1))

        tk.Label(main_frame, text="Дата заказа:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.entry_order_date = tk.Entry(main_frame)
        self.entry_order_date.grid(row=1, column=1, sticky="ew", pady=2)
        self.entry_order_date.insert(0, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        tk.Label(main_frame, text="Дата доставки:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.entry_delivery_date = tk.Entry(main_frame)
        self.entry_delivery_date.grid(row=2, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Адрес пункта выдачи:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.combo_address = ttk.Combobox(main_frame, values=[f"{a['id']}: {a['address']}" for a in self.addresses])
        self.combo_address.grid(row=3, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Клиент:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.combo_user = ttk.Combobox(main_frame, values=[f"{u['id']}: {u['full_name']}" for u in self.users])
        self.combo_user.grid(row=4, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Код для получения:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.entry_code = tk.Entry(main_frame)
        self.entry_code.grid(row=5, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Статус:").grid(row=6, column=0, sticky=tk.W, pady=2)
        self.combo_status = ttk.Combobox(main_frame, values=self.statuses)
        self.combo_status.grid(row=6, column=1, sticky="ew", pady=2)

        tk.Label(main_frame, text="Состав заказа (артикул, количество; через запятую):").grid(row=7, column=0, sticky=tk.W, pady=2)
        self.text_items = tk.Text(main_frame, height=5, width=40)
        self.text_items.grid(row=7, column=1, sticky="ew", pady=2)

        btn_frame = tk.Frame(main_frame)
        btn_frame.grid(row=8, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text="Сохранить", command=self.save_order, bg=COLOR_ACCENT).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Отмена", command=self.on_close, bg=COLOR_EXTRA_BG).pack(side=tk.LEFT, padx=5)

        main_frame.columnconfigure(1, weight=1)

    def load_order_data(self):
        cursor = self.app.db.conn.cursor()
        cursor.execute("""
            SELECT o.*, a.address, u.full_name
            FROM orders o
            LEFT JOIN addresses a ON o.address_id = a.id
            LEFT JOIN users u ON o.user_id = u.id
            WHERE o.id = ?
        """, (self.order_id,))
        order = cursor.fetchone()
        if order:
            self.entry_id.config(state=STATE_NORMAL)
            self.entry_id.delete(0, tk.END)
            self.entry_id.insert(0, order["id"])
            self.entry_id.config(state=STATE_READONLY)
            self.entry_order_date.delete(0, tk.END)
            self.entry_order_date.insert(0, order["order_date"] or "")
            self.entry_delivery_date.delete(0, tk.END)
            self.entry_delivery_date.insert(0, order["delivery_date"] or "")
            
            if order["address_id"]:
                self.combo_address.set(f"{order['address_id']}: {order['address']}")
            if order["user_id"]:
                self.combo_user.set(f"{order['user_id']}: {order['full_name']}")
                
            self.entry_code.delete(0, tk.END)
            self.entry_code.insert(0, order["pickup_code"] or "")
            self.combo_status.set(order["status"] or "")

            cursor.execute("SELECT product_article, quantity FROM order_items WHERE order_id=?", (self.order_id,))
            items = cursor.fetchall()
            items_str = ", ".join([f"{it['product_article']}, {it['quantity']}" for it in items])
            self.text_items.insert(TEXT_START, items_str)

    def _parse_order_items(self, items_str):
        items = []
        if not items_str:
            return items
            
        parts = [p.strip() for p in items_str.split(',')]
        if len(parts) % 2 != 0:
            raise ValueError("Нечетное количество элементов в составе заказа. Должны быть пары: артикул, количество")
            
        for i in range(0, len(parts), 2):
            article = parts[i]
            try:
                qty = int(parts[i+1])
            except ValueError:
                raise ValueError(f"Количество для артикула {article} должно быть целым числом")
            items.append((article, qty))
            
        return items

    def save_order(self):
        try:
            order_id = int(self.entry_id.get())
        except ValueError:
            messagebox.showerror(TITLE_ERROR, "Некорректный номер заказа")
            return
            
        order_date = self.entry_order_date.get().strip()
        delivery_date = self.entry_delivery_date.get().strip()
        address_text = self.combo_address.get().strip()
        user_text = self.combo_user.get().strip()
        code = self.entry_code.get().strip()
        status = self.combo_status.get().strip()
        items_str = self.text_items.get(TEXT_START, tk.END).strip()

        address_id = int(address_text.split(":")[0]) if address_text and ":" in address_text else None
        user_id = int(user_text.split(":")[0]) if user_text and ":" in user_text else None

        try:
            items = self._parse_order_items(items_str)
        except ValueError as e:
            messagebox.showerror(TITLE_ERROR, str(e))
            return

        cursor = self.app.db.conn.cursor()
        for article, _ in items:
            cursor.execute("SELECT article FROM products WHERE article=?", (article,))
            if not cursor.fetchone():
                messagebox.showerror(TITLE_ERROR, f"Товар с артикулом '{article}' не найден в базе")
                return

        cursor = self.app.db.conn.cursor()
        if self.order_id:
            cursor.execute('''
                UPDATE orders SET order_date=?, delivery_date=?, address_id=?, user_id=?, pickup_code=?, status=?
                WHERE id=?
            ''', (order_date, delivery_date, address_id, user_id, code, status, order_id))
            cursor.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
        else:
            cursor.execute('''
                INSERT INTO orders (id, order_date, delivery_date, address_id, user_id, pickup_code, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (order_id, order_date, delivery_date, address_id, user_id, code, status))

        for article, qty in items:
            cursor.execute("INSERT INTO order_items (order_id, product_article, quantity) VALUES (?, ?, ?)", (order_id, article, qty))

        self.app.db.conn.commit()
        self.parent.load_orders()
        self.on_close()

    def on_close(self):
        self.app.order_edit_window = None
        self.destroy()


# Запуск
if __name__ == "__main__":
    required_files = [
        "user_import.xlsx",
        "Tovar.xlsx",
        "Заказ_import.xlsx",
        "Пункты выдачи_import.xlsx",
        "picture.png",
    ]
    for f in required_files:
        if not os.path.exists(f):
            print(f"Предупреждение: файл {f} не найден")
    app = App()
    app.mainloop()